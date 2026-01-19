from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile
from fastapi.responses import StreamingResponse
from typing import List
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel

from app.db.supabase import testcases_collection, testcase_history_collection, projects_collection
from app.core.security import get_current_user_firestore
from app.core.permissions import check_write_permission
from app.schemas.testcase import TestCaseCreate, TestCaseUpdate, TestCase as TestCaseSchema
from app.services.ai_testcase_generator import generate_testcases_from_prd

router = APIRouter(redirect_slashes=False)


# Request/Response models for AI generation
class AIGenerateRequest(BaseModel):
    prd_content: str
    project_id: str


class AIGenerateResponse(BaseModel):
    testcases: List[dict]
    count: int


@router.post("", response_model=TestCaseSchema, status_code=status.HTTP_201_CREATED)
def create_testcase(
    testcase_in: TestCaseCreate,
    current_user: dict = Depends(get_current_user_firestore)
):
    # Check if user has permission to create test cases (viewer and developer cannot create)
    check_write_permission(current_user, "í…ŒìŠ¤íŠ¸ì¼€ì´ìŠ¤")

    # Convert Pydantic model to dict using json-compatible way to handle Enums
    import json
    testcase_data = json.loads(testcase_in.json())
    
    # Debug logging
    print(f"ğŸ“ Creating testcase - Input tags: {testcase_data.get('tags')} (type: {type(testcase_data.get('tags'))})")
    
    # Add created_by
    testcase_data['created_by'] = current_user.get('id')
    
    # Ensure tags is a list (handle legacy string format)
    if testcase_data.get('tags') is None:
        testcase_data['tags'] = []
    elif isinstance(testcase_data.get('tags'), str):
        # Convert comma-separated string to list
        print(f"âš ï¸ Converting string tags to list: {testcase_data['tags']}")
        testcase_data['tags'] = [t.strip() for t in testcase_data['tags'].split(',') if t.strip()]

    print(f"ğŸ“ Final testcase data tags: {testcase_data['tags']}")
    
    try:
        testcase = testcases_collection.create(testcase_data)
    except Exception as e:
        # Handle missing created_by column (Schema mismatch)
        error_str = str(e)
        if "Could not find the 'created_by' column" in error_str:
            print("âš ï¸ 'created_by' column missing in database schema. Retrying without it.")
            if 'created_by' in testcase_data:
                del testcase_data['created_by']
            testcase = testcases_collection.create(testcase_data)
        else:
            raise e
            
    return testcase


@router.get("", response_model=List[TestCaseSchema])
def list_testcases(
    project_id: str = None,
    skip: int = 0,
    limit: int = 100,
    current_user: dict = Depends(get_current_user_firestore)
):
    if project_id:
        testcases = testcases_collection.query('project_id', '==', project_id)
    else:
        testcases = testcases_collection.list(limit=limit)

    return testcases[skip:skip+limit]


@router.get("/{testcase_id}", response_model=TestCaseSchema)
def get_testcase(
    testcase_id: str,
    current_user: dict = Depends(get_current_user_firestore)
):
    testcase = testcases_collection.get(testcase_id)
    if not testcase:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test case not found"
        )
    return testcase


@router.put("/{testcase_id}", response_model=TestCaseSchema)
def update_testcase(
    testcase_id: str,
    testcase_in: TestCaseUpdate,
    current_user: dict = Depends(get_current_user_firestore)
):
    # Check if user has permission to update test cases (viewer and developer cannot update)
    check_write_permission(current_user, "í…ŒìŠ¤íŠ¸ì¼€ì´ìŠ¤")

    testcase = testcases_collection.get(testcase_id)
    if not testcase:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test case not found"
        )

    # Get current version number
    history_records = testcase_history_collection.query('testcase_id', '==', testcase_id)
    new_version = len(history_records) + 1

    # Save current state to history before updating
    history_data = {
        'testcase_id': testcase_id,
        'version': new_version,
        'title': testcase.get('title'),
        'description': testcase.get('description'),
        'preconditions': testcase.get('preconditions'),
        'steps': testcase.get('steps'),
        'expected_result': testcase.get('expected_result'),
        'priority': testcase.get('priority'),
        'test_type': testcase.get('test_type'),
        'modified_by': current_user['id'],  # Supabase uses modified_by instead of changed_by
        'change_note': testcase_in.dict().get('change_note', None)  # Pydantic v1 uses .dict()
    }
    testcase_history_collection.create(history_data)

    # Update testcase
    update_data = testcase_in.dict(exclude_unset=True)  # Pydantic v1 uses .dict()
    if 'change_note' in update_data:
        del update_data['change_note']

    testcases_collection.update(testcase_id, update_data)

    # Return updated testcase
    updated_testcase = testcases_collection.get(testcase_id)
    return updated_testcase


@router.delete("/{testcase_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_testcase(
    testcase_id: str,
    current_user: dict = Depends(get_current_user_firestore)
):
    # Check if user has permission to delete test cases (viewer and developer cannot delete)
    check_write_permission(current_user, "í…ŒìŠ¤íŠ¸ì¼€ì´ìŠ¤")

    testcase = testcases_collection.get(testcase_id)
    if not testcase:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test case not found"
        )

    testcases_collection.delete(testcase_id)
    return None


@router.get("/{testcase_id}/history")
def get_testcase_history(
    testcase_id: str,
    current_user: dict = Depends(get_current_user_firestore)
):
    testcase = testcases_collection.get(testcase_id)
    if not testcase:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Test case not found"
        )

    history_records = testcase_history_collection.query('testcase_id', '==', testcase_id)
    # Sort by version descending
    history_records.sort(key=lambda x: x.get('version', 0), reverse=True)

    return history_records


@router.get("/template/download")
def download_template(
    current_user: dict = Depends(get_current_user_firestore)
):
    """Download Excel template for test case import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "í…ŒìŠ¤íŠ¸ ì¼€ì´ìŠ¤"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define headers
    headers = [
        "í”„ë¡œì íŠ¸ ì´ë¦„*",
        "ì œëª©*",
        "ì„¤ëª…",
        "ì‚¬ì „ì¡°ê±´",
        "ìˆ˜í–‰ë°©ë²•*",
        "ì˜ˆìƒê²°ê³¼*",
        "ìš°ì„ ìˆœìœ„*",
        "í…ŒìŠ¤íŠ¸ ìœ í˜•*"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add example row
    example_row = [
        "ìƒ˜í”Œ í”„ë¡œì íŠ¸",
        "ë¡œê·¸ì¸ ê¸°ëŠ¥ í…ŒìŠ¤íŠ¸",
        "ì‚¬ìš©ìê°€ ì´ë©”ì¼ê³¼ ë¹„ë°€ë²ˆí˜¸ë¡œ ë¡œê·¸ì¸í•  ìˆ˜ ìˆëŠ”ì§€ í™•ì¸",
        "1. ì‚¬ìš©ì ê³„ì •ì´ ìƒì„±ë˜ì–´ ìˆì–´ì•¼ í•¨\n2. ë¡œê·¸ì¸ í˜ì´ì§€ì— ì ‘ê·¼ ê°€ëŠ¥í•´ì•¼ í•¨",
        "1. ë¡œê·¸ì¸ í˜ì´ì§€ë¡œ ì´ë™\n2. ì´ë©”ì¼ ì…ë ¥\n3. ë¹„ë°€ë²ˆí˜¸ ì…ë ¥\n4. ë¡œê·¸ì¸ ë²„íŠ¼ í´ë¦­",
        "ì‚¬ìš©ìê°€ ì„±ê³µì ìœ¼ë¡œ ë¡œê·¸ì¸ë˜ê³  ëŒ€ì‹œë³´ë“œ í˜ì´ì§€ë¡œ ì´ë™ë¨",
        "high",
        "functional"
    ]

    for col_num, value in enumerate(example_row, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Add instruction sheet
    ws_info = wb.create_sheet("ì‘ì„± ê°€ì´ë“œ")
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 50

    instructions = [
        ("í•„ìˆ˜ í•„ë“œ", "* í‘œì‹œê°€ ìˆëŠ” í•„ë“œëŠ” ë°˜ë“œì‹œ ì…ë ¥í•´ì•¼ í•©ë‹ˆë‹¤"),
        ("í”„ë¡œì íŠ¸ ì´ë¦„", "í…ŒìŠ¤íŠ¸ ì¼€ì´ìŠ¤ê°€ ì†í•  í”„ë¡œì íŠ¸ì˜ ì´ë¦„ì„ ì…ë ¥í•©ë‹ˆë‹¤ (ì •í™•íˆ ì¼ì¹˜í•´ì•¼ í•¨)"),
        ("ì œëª©", "í…ŒìŠ¤íŠ¸ ì¼€ì´ìŠ¤ì˜ ì œëª©ì„ ì…ë ¥í•©ë‹ˆë‹¤"),
        ("ì„¤ëª…", "í…ŒìŠ¤íŠ¸ ì¼€ì´ìŠ¤ì— ëŒ€í•œ ìƒì„¸ ì„¤ëª…ì„ ì…ë ¥í•©ë‹ˆë‹¤ (ì„ íƒ)"),
        ("ì‚¬ì „ì¡°ê±´", "í…ŒìŠ¤íŠ¸ ìˆ˜í–‰ ì „ ì¤€ë¹„í•´ì•¼ í•  ì¡°ê±´ì„ ì…ë ¥í•©ë‹ˆë‹¤ (ì„ íƒ)"),
        ("ìˆ˜í–‰ë°©ë²•", "í…ŒìŠ¤íŠ¸ë¥¼ ìˆ˜í–‰í•˜ëŠ” ë‹¨ê³„ë³„ ì ˆì°¨ë¥¼ ì…ë ¥í•©ë‹ˆë‹¤"),
        ("ì˜ˆìƒê²°ê³¼", "í…ŒìŠ¤íŠ¸ ìˆ˜í–‰ í›„ ì˜ˆìƒë˜ëŠ” ê²°ê³¼ë¥¼ ì…ë ¥í•©ë‹ˆë‹¤"),
        ("ìš°ì„ ìˆœìœ„", "high, medium, low ì¤‘ í•˜ë‚˜ë¥¼ ì…ë ¥í•©ë‹ˆë‹¤"),
        ("í…ŒìŠ¤íŠ¸ ìœ í˜•", "functional, performance, security, usability ì¤‘ í•˜ë‚˜ë¥¼ ì…ë ¥í•©ë‹ˆë‹¤"),
    ]

    for row_num, (field, desc) in enumerate(instructions, 1):
        ws_info.cell(row=row_num, column=1, value=field).font = Font(bold=True)
        ws_info.cell(row=row_num, column=2, value=desc)

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=testcase_template.xlsx"}
    )


@router.post("/import/excel")
async def import_excel(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user_firestore)
):
    """Import test cases from Excel file"""
    # Check if user has permission to import test cases (viewer and developer cannot import)
    check_write_permission(current_user, "í…ŒìŠ¤íŠ¸ì¼€ì´ìŠ¤")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are allowed"
        )

    try:
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active

        # Skip header row
        imported_count = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Skip empty rows
                continue

            try:
                project_name, title, description, preconditions, steps, expected_result, priority, test_type = row[:8]

                # Validate required fields
                if not all([project_name, title, steps, expected_result, priority, test_type]):
                    errors.append(f"í–‰ {row_num}: í•„ìˆ˜ í•„ë“œê°€ ëˆ„ë½ë˜ì—ˆìŠµë‹ˆë‹¤")
                    continue

                # Find project by name
                projects = projects_collection.query('name', '==', str(project_name))
                if not projects:
                    errors.append(f"í–‰ {row_num}: í”„ë¡œì íŠ¸ '{project_name}'ì„(ë¥¼) ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤")
                    continue

                project = projects[0]
                project_id = project['id']

                # Validate priority
                if priority not in ['high', 'medium', 'low']:
                    errors.append(f"í–‰ {row_num}: ìš°ì„ ìˆœìœ„ëŠ” high, medium, low ì¤‘ í•˜ë‚˜ì—¬ì•¼ í•©ë‹ˆë‹¤")
                    continue

                # Validate test_type
                if test_type not in ['functional', 'regression', 'smoke', 'integration', 'performance', 'security']:
                    errors.append(f"í–‰ {row_num}: í…ŒìŠ¤íŠ¸ ìœ í˜•ì€ functional, regression, smoke, integration, performance, security ì¤‘ í•˜ë‚˜ì—¬ì•¼ í•©ë‹ˆë‹¤")
                    continue

                # Create test case
                testcase_data = {
                    'project_id': str(project_id),
                    'title': str(title),
                    'description': str(description) if description else None,
                    'preconditions': str(preconditions) if preconditions else None,
                    'steps': str(steps),
                    'expected_result': str(expected_result),
                    'priority': str(priority),
                    'test_type': str(test_type)
                }

                testcases_collection.create(testcase_data)
                imported_count += 1

            except Exception as e:
                errors.append(f"í–‰ {row_num}: {str(e)}")

        return {
            "imported_count": imported_count,
            "errors": errors,
            "success": imported_count > 0
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel íŒŒì¼ ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}"
        )


@router.post("/ai/generate", response_model=AIGenerateResponse)
def generate_testcases_with_ai(
    request: AIGenerateRequest,
    current_user: dict = Depends(get_current_user_firestore)
):
    """Generate test cases from PRD content using AI"""
    # Check if user has permission to create test cases
    check_write_permission(current_user, "í…ŒìŠ¤íŠ¸ì¼€ì´ìŠ¤")

    # Validate project exists
    project = projects_collection.get(request.project_id)
    if not project:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )

    try:
        # Generate test cases using AI
        generated_testcases = generate_testcases_from_prd(
            prd_content=request.prd_content,
            project_name=project.get('name', '')
        )

        # Return generated test cases (not saved to database yet - user can review and edit)
        return AIGenerateResponse(
            testcases=generated_testcases,
            count=len(generated_testcases)
        )

    except ValueError as e:
        # Specific error from AI service (API key missing, parsing error, etc.)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI í…ŒìŠ¤íŠ¸ ì¼€ì´ìŠ¤ ìƒì„± ì¤‘ ì˜¤ë¥˜ê°€ ë°œìƒí–ˆìŠµë‹ˆë‹¤: {str(e)}"
        )
